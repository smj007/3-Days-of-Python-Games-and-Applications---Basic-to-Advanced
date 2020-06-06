import openpyxl as xl

workbook = xl.load_workbook("sample1 - data for automation with python.xlsx")
sheet = workbook["Sheet1"]

for row in range(2, sheet.max_row + 1):
    physics = sheet.cell(row, 3).value
    maths = sheet.cell(row, 4).value
    history = sheet.cell(row, 5).value
    geography = sheet.cell(row, 6).value
    biology = sheet.cell(row, 7).value
    chemistry = sheet.cell(row, 8).value
    total_marks = physics + maths + history + geography + biology + chemistry
    total_marks_cell = sheet.cell(row, 9)
    total_marks_cell.value = total_marks

    total_percentages = total_marks // 6
    total_percentages_cell = sheet.cell(row, 10)
    total_percentages_cell.value = total_percentages

workbook.save("Final-results.xlsx")
print("Document saved successfully")
